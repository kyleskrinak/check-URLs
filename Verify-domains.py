# Define the following in config.py
# EXCEL_FILE_PATH = "path/to/your/excel/file.xlsx"
# WORKSHEETS_TO_CHECK = ["Sheet1", "Sheet2", "Sheet3"]  # List of worksheet names to check
# DOMAIN_COLUMN = 1  # Column number containing the domain (1 for column A, 2 for column B, and so on)
# LOG_FILE_PATH = "path/to/your/log/file.log"  # Path to the log file
# TIMEOUT = 10  # Adjust the timeout value in seconds as needed

import openpyxl
import requests
import logging
import concurrent.futures
from config import WORKSHEETS_TO_CHECK, EXCEL_FILE_PATH, DOMAIN_COLUMN, LOG_FILE_PATH, TIMEOUT

def check_status_code(url):
    try:
        with requests.get(url, timeout=TIMEOUT) as response:
            return response.status_code
    except requests.exceptions.RequestException:
        return None

def process_url(domain, logger):
    url = f"https://{domain}"
    status_code = check_status_code(url)

    if status_code is None:
        logger.error(f"Error: {url} - Unable to connect")
    elif status_code == 200:
        logger.info(f"Success: {url} - 200 OK")
    else:
        logger.warning(f"Warning: {url} - Status Code: {status_code}")

def process_worksheet(worksheet, domain_column, logger):
    row_start = 2
    domains = [row[domain_column - 1] for row in worksheet.iter_rows(min_row=row_start, values_only=True)]

    empty_cell_count = 0
    with concurrent.futures.ThreadPoolExecutor() as executor:
        for domain in domains:
            if not domain:
                empty_cell_count += 1
                if empty_cell_count >= 5:
                    logger.warning("Five consecutive empty cells encountered. Starting a new loop for the next worksheet.")
                    return False
                continue

            print(f"Checking domain: {domain}")
            process_url(domain, logger)
            empty_cell_count = 0

    logger.info("Worksheet processed successfully.")
    return True

def main():
    logging.basicConfig(filename=LOG_FILE_PATH, level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    logger = logging.getLogger()

    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)

    for worksheet_name in WORKSHEETS_TO_CHECK:
        if worksheet_name not in workbook.sheetnames:
            logger.warning(f"Worksheet '{worksheet_name}' not found.")
            continue

        worksheet = workbook[worksheet_name]
        logger.info(f"Processing worksheet: {worksheet_name}")
        while process_worksheet(worksheet, DOMAIN_COLUMN, logger):
            continue

    workbook.close()

if __name__ == "__main__":
    main()